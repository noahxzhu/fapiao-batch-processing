import os
import re

import pandas as pd
import pymupdf
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm

fapiao_dir = "/Users/noah/Downloads/invoice"
export_excel_file_name = "invoices.xlsx"


def list_files(directory):
    entries = os.listdir(directory)
    files = [file for file in entries if file.endswith(".pdf")]
    return files


def is_fapiao_code(s, is_zzs=False):
    pattern = r"^\d{8}$" if is_zzs else r"^\d{20}$"
    if re.match(pattern, s):
        return True
    else:
        return False


def is_valid_date(date_str):
    pattern = r"^\d{4}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])$"
    return bool(re.match(pattern, date_str))


files = list_files(fapiao_dir)

data = []

for filename in tqdm(files):
    file_path = f"{fapiao_dir}/{filename}"
    is_zzs = False
    fapiao_code = ""
    price = 0
    date = ""

    with pymupdf.open(file_path) as doc:
        page = doc.load_page(0)
        blocks = page.get_textpage().extractBLOCKS()
        for block in blocks:
            block_text = block[4]
            if "机器编号" in block_text:
                is_zzs = True

        for block in blocks:
            block_text = block[4].strip()
            splits = block_text.split("\n")
            for s in splits:
                if is_fapiao_code(s, is_zzs):
                    fapiao_code = s

            if "¥" in block_text:
                price_texts = block_text.split("\n")
                for price_text in price_texts:
                    if "¥" in price_text:
                        index = price_text.find("¥") + 1
                        new_price = float(price_text[index:])
                        if new_price > price:
                            price = new_price

            if "￥" in block_text:
                price_texts = block_text.split("\n")
                for price_text in price_texts:
                    if "￥" in price_text:
                        index = price_text.find("￥") + 1
                        new_price = float(price_text[index:])
                        if new_price > price:
                            price = new_price

            date_str = (
                block_text.replace("\n", "")
                .replace("年", "")
                .replace("月", "")
                .replace("日", "")
                .replace(" ", "")
            )
            if is_valid_date(date_str):
                date = date_str

    new_filename = f"{fapiao_code}+{price}.pdf"
    new_file_path = f"{fapiao_dir}/{new_filename}"
    try:
        os.rename(file_path, new_file_path)
    except Exception:
        print(f"rename {filename} to {new_filename} failed")

    data.append([new_filename, fapiao_code, date, price])

total_price = 0
for d in data:
    total_price = total_price + d[3]

data.append(["", "", "", total_price])
df = pd.DataFrame(data, columns=["文件名", "发票编号", "开票日期", "金额"])

excel_path = f"{fapiao_dir}/{export_excel_file_name}"
if os.path.isfile(excel_path):
    os.remove(excel_path)

df.to_excel(excel_path, index=False)


# Adjust the column width
wb = load_workbook(excel_path)
ws = wb.active

for col in ws.columns:
    max_length = 0
    column = col[0].column

    for cell in col:
        if cell.value:
            cell_length = len(str(cell.value))
            max_length = max(max_length, cell_length)

    adjusted_width = max_length + 2
    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

wb.save(excel_path)
